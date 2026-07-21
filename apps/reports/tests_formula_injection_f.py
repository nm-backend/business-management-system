"""
Этап F — регрессия Formula Injection (CSV/Excel injection) в xlsx-экспорте.

Значение, начинающееся с = + - @ / tab / CR, Excel исполняет как формулу.
Пользовательские данные (имена клиентов/материалов, комментарии) попадают в
отчёты, поэтому такой ввод должен нейтрализоваться префиксом апострофа.
"""
from io import BytesIO

from django.test import TestCase
from openpyxl import load_workbook

from apps.reports.views import xlsx_response, _sanitize_xlsx_cell


class FormulaInjectionTests(TestCase):
    def _cells(self, rows):
        resp = xlsx_response(rows, 'x.xlsx', 'Sheet')
        wb = load_workbook(BytesIO(resp.content))
        return wb.active

    def test_formula_cells_are_neutralized(self):
        ws = self._cells([
            ['Название', 'Кол-во'],
            ['=1+2', 5],
            ['@SUM(A1:A9)', 3],
            ['+cmd', 1],
            ['-2+3', 2],
            ['обычный', 9],
        ])
        self.assertEqual(ws['A2'].value, "'=1+2")
        self.assertEqual(ws['A3'].value, "'@SUM(A1:A9)")
        self.assertEqual(ws['A4'].value, "'+cmd")
        self.assertEqual(ws['A5'].value, "'-2+3")
        self.assertEqual(ws['A6'].value, 'обычный')       # безопасное не тронуто

    def test_numbers_untouched(self):
        # Числовые значения не должны превращаться в текст.
        self.assertEqual(_sanitize_xlsx_cell(5), 5)
        self.assertEqual(_sanitize_xlsx_cell(-5), -5)
        self.assertEqual(_sanitize_xlsx_cell('обычный текст'), 'обычный текст')
