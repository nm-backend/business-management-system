"""
Report generation services for PDF and Excel.

Содержит функции для генерации отчётов:
- Для владельца: полный финансовый отчёт (выручка, прибыль, расходы, касса, долги)
- Для администратора: операционный отчёт (заказы, склад, работники)

Формат: PDF (ReportLab) и Excel (openpyxl).
"""
import logging
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

logger = logging.getLogger(__name__)


# ──────────────────────────────────────────────
#  ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ──────────────────────────────────────────────

def _get_period_range(period, custom_start=None, custom_end=None):
    """Возвращает (date_from, date_to) для заданного периода."""
    today = date.today()
    if period == 'today':
        return today, today
    elif period == 'yesterday':
        y = today - timedelta(days=1)
        return y, y
    elif period == 'week':
        start = today - timedelta(days=today.weekday())
        return start, today
    elif period == 'month':
        return today.replace(day=1), today
    elif period == 'quarter':
        qm = ((today.month - 1) // 3) * 3 + 1
        return today.replace(month=qm, day=1), today
    elif period == 'year':
        return today.replace(month=1, day=1), today
    elif period == 'custom' and custom_start and custom_end:
        from datetime import datetime as dt
        try:
            return dt.strptime(custom_start, '%Y-%m-%d').date(), dt.strptime(custom_end, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return today.replace(day=1), today
    return today.replace(day=1), today


def _fmt(n):
    """Форматирует Decimal/int/str в читаемую строку с пробелами между разрядами."""
    if n is None:
        return '0'
    try:
        val = float(str(n))
        return f"{val:,.0f}".replace(',', ' ')
    except (ValueError, TypeError):
        return str(n)


# ──────────────────────────────────────────────
#  СБОР ДАННЫХ
# ──────────────────────────────────────────────

def _collect_financial_data(period, custom_start=None, custom_end=None):
    """
    Собирает полные финансовые данные для owner-отчёта.
    Возвращает dict со всеми показателями.
    """
    date_from, date_to = _get_period_range(period, custom_start, custom_end)
    data = {
        'period_key': period,
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        'title': f"Финансовый отчёт: {date_from} — {date_to}",
    }

    try:
        from apps.finance.models import Expense, ExpenseCategory, WorkerPayment
        from apps.orders.models import Order, PaymentStatus
        from django.db.models import Sum, Q

        # Выручка
        delivered = Order.objects.filter(
            status='delivered',
            updated_at__date__gte=date_from,
            updated_at__date__lte=date_to,
        )
        rev = delivered.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
        data['revenue'] = _fmt(rev)

        # Себестоимость
        cog = Decimal('0')
        for o in delivered.select_related('product'):
            if o.product and o.product.cost_price:
                cog += o.product.cost_price * o.quantity
        data['cost_of_goods'] = _fmt(cog)

        # Расходы
        expenses = Expense.objects.filter(date__gte=date_from, date__lte=date_to)
        exp_total = expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        data['expenses_total'] = _fmt(exp_total)

        # Налоги
        taxes = expenses.filter(category=ExpenseCategory.TAXES)
        tax_total = taxes.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        data['taxes'] = _fmt(tax_total)

        # Потери
        losses = expenses.filter(category__in=[ExpenseCategory.MATERIAL_LOSS, ExpenseCategory.DEFECT])
        loss_total = losses.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        data['losses'] = _fmt(loss_total)

        # Зарплаты
        salaries = WorkerPayment.objects.filter(payment_date__gte=date_from, payment_date__lte=date_to)
        sal_total = salaries.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        data['salaries'] = _fmt(sal_total)

        # Вывод владельца
        owner_wd = expenses.filter(category=ExpenseCategory.OWNER_WITHDRAWAL)
        ow_total = owner_wd.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        data['owner_withdrawal'] = _fmt(ow_total)

        # Валовая прибыль
        gross = rev - cog
        data['gross_profit'] = _fmt(gross)

        # Чистая прибыль
        net = rev - cog - sal_total - exp_total - tax_total - loss_total
        data['net_profit'] = _fmt(net)

        # Касса
        all_paid = Order.objects.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
        cash = all_paid - exp_total - sal_total - ow_total
        data['cash_in_register'] = _fmt(cash)

        # Долги клиентов
        debts = Order.objects.filter(~Q(payment_status=PaymentStatus.PAID))
        from django.db.models import F
        client_debt = sum(
            (item.total_amount - item.paid_amount for item in debts),
            Decimal('0')
        )
        data['client_debts'] = _fmt(client_debt)

        # Долги работников (расходы с категорией worker_debt)
        worker_debt_exp = expenses.filter(category=ExpenseCategory.WORKER_DEBT)
        wd_total = worker_debt_exp.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        data['worker_debts'] = _fmt(wd_total)

        # Кол-во заказов
        data['orders_count'] = delivered.count()
        data['total_orders'] = Order.objects.filter(
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
        ).count()

        # Категории расходов
        cat_data = {}
        for expense in expenses.values('category').annotate(total=Sum('amount')):
            cat_data[expense['category']] = _fmt(expense['total'])
        data['expense_categories'] = cat_data

    except Exception as e:
        logger.error(f"Financial data collection error: {e}")
        for k in ['revenue', 'cost_of_goods', 'expenses_total', 'taxes', 'losses',
                   'salaries', 'owner_withdrawal', 'gross_profit', 'net_profit',
                   'cash_in_register', 'client_debts', 'worker_debts']:
            data.setdefault(k, '0')
        data.setdefault('orders_count', 0)
        data.setdefault('total_orders', 0)
        data.setdefault('expense_categories', {})

    return data


def _collect_operational_data(period, custom_start=None, custom_end=None):
    """
    Собирает операционные данные для admin-отчёта (без финансов).
    """
    date_from, date_to = _get_period_range(period, custom_start, custom_end)
    data = {
        'period_key': period,
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        'title': f"Операционный отчёт: {date_from} — {date_to}",
    }

    try:
        from apps.orders.models import Order, OrderStatus
        from apps.production.models import Task, WorkRecord
        from apps.warehouse.models import RawMaterial, FinishedProduct
        from django.db.models import Count, F, Q, Sum

        # Заказы
        period_orders = Order.objects.filter(
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
        )
        data['new_orders'] = period_orders.filter(status=OrderStatus.NEW).count()
        data['in_progress_orders'] = period_orders.filter(
            status__in=[OrderStatus.IN_PROGRESS, OrderStatus.SENT_TO_WORKER,
                        OrderStatus.ACCEPTED_BY_WORKER]
        ).count()
        data['ready_orders'] = period_orders.filter(status=OrderStatus.READY).count()
        data['delivered_orders'] = period_orders.filter(status=OrderStatus.DELIVERED).count()
        data['overdue_orders'] = period_orders.filter(is_overdue=True).count()
        data['total_orders'] = period_orders.count()

        # Неоплаченные (без суммы — только кол-во для admin)
        from apps.orders.models import PaymentStatus
        data['unpaid_orders_count'] = period_orders.filter(
            ~Q(payment_status=PaymentStatus.PAID)
        ).count()

        # Склад
        data['total_materials'] = RawMaterial.objects.filter(is_archived=False).count()
        data['low_stock_materials'] = RawMaterial.objects.filter(
            is_archived=False, quantity__lte=F('min_stock')
        ).count()
        data['total_products'] = FinishedProduct.objects.filter(is_archived=False).count()

        # Задачи
        period_tasks = Task.objects.filter(created_at__date__gte=date_from, created_at__date__lte=date_to)
        data['tasks_created'] = period_tasks.count()
        data['tasks_completed'] = period_tasks.filter(status='confirmed').count()

        # Работники (кто сколько сделал — без сумм)
        worker_qs = WorkRecord.objects.filter(
            status='confirmed',
            confirmed_at__date__gte=date_from,
            confirmed_at__date__lte=date_to,
        ).values('worker__username').annotate(
            works=Count('id'),
            total_qty=Sum('quantity')
        ).order_by('-works')[:10]
        data['worker_stats'] = [
            {'name': w['worker__username'], 'works': w['works'], 'quantity': str(w['total_qty'])}
            for w in worker_qs
        ]

    except Exception as e:
        logger.error(f"Operational data collection error: {e}")
        for k in ['new_orders', 'in_progress_orders', 'ready_orders', 'delivered_orders',
                   'overdue_orders', 'total_orders', 'unpaid_orders_count',
                   'total_materials', 'low_stock_materials', 'total_products',
                   'tasks_created', 'tasks_completed']:
            data.setdefault(k, 0)
        data.setdefault('worker_stats', [])

    return data


# ──────────────────────────────────────────────
#  ГЕНЕРАЦИЯ PDF (ReportLab)
# ──────────────────────────────────────────────

def generate_financial_pdf(period='month', custom_start=None, custom_end=None):
    """
    Генерирует PDF с полным финансовым отчётом для владельца.

    Возвращает:
        BytesIO — содержимое PDF файла
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    )

    data = _collect_financial_data(period, custom_start, custom_end)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title=data['title'])
    styles = getSampleStyleSheet()
    elements = []

    # Стили
    title_style = ParagraphStyle(
        'ReportTitle', parent=styles['Title'],
        fontSize=18, spaceAfter=12, alignment=1,  # center
    )
    subtitle_style = ParagraphStyle(
        'Subtitle', parent=styles['Normal'],
        fontSize=10, textColor=colors.grey, alignment=1, spaceAfter=20,
    )
    section_style = ParagraphStyle(
        'Section', parent=styles['Heading2'],
        fontSize=14, spaceBefore=16, spaceAfter=8,
    )
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=10)

    # Заголовок
    elements.append(Paragraph(data['title'], title_style))
    if data['period_key'] != 'custom':
        elements.append(Paragraph(
            f"Период: {data['date_from']} — {data['date_to']}", subtitle_style
        ))
    elements.append(Spacer(1, 12))

    # Финансовые показатели
    elements.append(Paragraph("Основные показатели", section_style))
    fin_rows = [
        ['Показатель', 'Значение (сум)'],
        ['Выручка', data['revenue']],
        ['Себестоимость', data['cost_of_goods']],
        ['Валовая прибыль', data['gross_profit']],
        ['Расходы', data['expenses_total']],
        ['Налоги', data['taxes']],
        ['Потери', data['losses']],
        ['Зарплаты', data['salaries']],
        ['Вывод владельца', data['owner_withdrawal']],
        ['Чистая прибыль', data['net_profit']],
        ['Касса', data['cash_in_register']],
        ['Долги клиентов', data['client_debts']],
        ['Долги работников', data['worker_debts']],
    ]
    fin_table = Table(fin_rows, colWidths=[250, 150])
    fin_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    # Выделить чистую прибыль жирным
    fin_table.setStyle(TableStyle([
        ('FONTNAME', (0, 9), (-1, 9), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 9), (-1, 9), colors.HexColor('#27ae60')),
    ]))
    elements.append(fin_table)
    elements.append(Spacer(1, 16))

    # Статистика заказов
    elements.append(Paragraph("Статистика заказов", section_style))
    order_rows = [
        ['Показатель', 'Значение'],
        ['Всего заказов за период', str(data['total_orders'])],
        ['Выполнено (доставлено)', str(data['orders_count'])],
    ]
    order_table = Table(order_rows, colWidths=[250, 150])
    order_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    elements.append(order_table)

    # Расходы по категориям
    if data['expense_categories']:
        elements.append(Spacer(1, 16))
        elements.append(Paragraph("Расходы по категориям", section_style))
        cat_rows = [['Категория', 'Сумма']]
        for cat, total in sorted(data['expense_categories'].items()):
            cat_rows.append([cat, total])
        cat_table = Table(cat_rows, colWidths=[250, 150])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        elements.append(cat_table)

    # Подвал
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(
        f"Сгенерировано SkladPro.Nod • {date.today().isoformat()}",
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=1)
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_operational_pdf(period='month', custom_start=None, custom_end=None):
    """
    Генерирует PDF с операционным отчётом для администратора (без финансов).
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    )

    data = _collect_operational_data(period, custom_start, custom_end)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title=data['title'])
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle(
        'ReportTitle', parent=styles['Title'],
        fontSize=18, spaceAfter=12, alignment=1,
    )
    subtitle_style = ParagraphStyle(
        'Subtitle', parent=styles['Normal'],
        fontSize=10, textColor=colors.grey, alignment=1, spaceAfter=20,
    )
    section_style = ParagraphStyle(
        'Section', parent=styles['Heading2'],
        fontSize=14, spaceBefore=16, spaceAfter=8,
    )

    elements.append(Paragraph(data['title'], title_style))
    elements.append(Paragraph(
        f"Период: {data['date_from']} — {data['date_to']}", subtitle_style
    ))
    elements.append(Spacer(1, 12))

    # Заказы
    elements.append(Paragraph("Заказы", section_style))
    order_rows = [
        ['Статус', 'Количество'],
        ['Новые', str(data['new_orders'])],
        ['В работе', str(data['in_progress_orders'])],
        ['Готовые', str(data['ready_orders'])],
        ['Доставленные', str(data['delivered_orders'])],
        ['Просроченные', str(data['overdue_orders'])],
        ['Неоплаченные (заказы)', str(data['unpaid_orders_count'])],
        ['Всего за период', str(data['total_orders'])],
    ]
    t = Table(order_rows, colWidths=[250, 150])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 12))

    # Склад
    elements.append(Paragraph("Склад", section_style))
    wh_rows = [
        ['Показатель', 'Значение'],
        ['Материалов на складе', str(data['total_materials'])],
        ['С низким остатком', str(data['low_stock_materials'])],
        ['Готовой продукции', str(data['total_products'])],
    ]
    t = Table(wh_rows, colWidths=[250, 150])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 12))

    # Задачи
    elements.append(Paragraph("Задачи", section_style))
    task_rows = [
        ['Показатель', 'Значение'],
        ['Создано задач', str(data['tasks_created'])],
        ['Выполнено', str(data['tasks_completed'])],
        ['Выполняемость', f"{round(data['tasks_completed'] / max(data['tasks_created'], 1) * 100)}%"],
    ]
    t = Table(task_rows, colWidths=[250, 150])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    elements.append(t)

    # Активность работников
    if data.get('worker_stats'):
        elements.append(Spacer(1, 12))
        elements.append(Paragraph("Активность работников", section_style))
        wr_rows = [['Работник', 'Выполнено работ', 'Общее количество']]
        for w in data['worker_stats']:
            wr_rows.append([w['name'], str(w['works']), w['quantity']])
        t = Table(wr_rows, colWidths=[150, 130, 120])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        elements.append(t)

    # Подвал
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(
        f"Сгенерировано SkladPro.Nod • {date.today().isoformat()}",
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=1)
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ──────────────────────────────────────────────
#  ГЕНЕРАЦИЯ EXCEL (openpyxl)
# ──────────────────────────────────────────────

def generate_financial_excel(period='month', custom_start=None, custom_end=None):
    """
    Генерирует Excel с полным финансовым отчётом для владельца.

    Возвращает:
        BytesIO — содержимое Excel файла
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = _collect_financial_data(period, custom_start, custom_end)
    wb = Workbook()

    # ── Лист 1: Основные показатели ──
    ws = wb.active
    ws.title = "Финансы"
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    alt_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )

    ws.merge_cells('A1:B1')
    ws['A1'] = data['title']
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Показатель', 'Значение (сум)']
    metrics = [
        ('Выручка', data['revenue']),
        ('Себестоимость', data['cost_of_goods']),
        ('Валовая прибыль', data['gross_profit']),
        ('Расходы', data['expenses_total']),
        ('Налоги', data['taxes']),
        ('Потери', data['losses']),
        ('Зарплаты', data['salaries']),
        ('Вывод владельца', data['owner_withdrawal']),
        ('Чистая прибыль', data['net_profit']),
        ('Касса', data['cash_in_register']),
        ('Долги клиентов', data['client_debts']),
        ('Долги работников', data['worker_debts']),
        ('', ''),
        ('Заказов за период', str(data['total_orders'])),
        ('Доставлено', str(data['orders_count'])),
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_idx, (label, value) in enumerate(metrics, 4):
        ws.cell(row=row_idx, column=1, value=label).border = thin_border
        val_cell = ws.cell(row=row_idx, column=2, value=value)
        val_cell.border = thin_border
        val_cell.alignment = Alignment(horizontal='right')
        if row_idx % 2 == 0:
            ws.cell(row=row_idx, column=1).fill = alt_fill
            val_cell.fill = alt_fill

    # Выделить чистую прибыль
    profit_row = 12  # Чистая прибыль
    ws.cell(row=profit_row, column=1).font = Font(bold=True, color='27AE60')
    ws.cell(row=profit_row, column=2).font = Font(bold=True, color='27AE60')

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

    # ── Лист 2: Расходы по категориям ──
    if data['expense_categories']:
        ws2 = wb.create_sheet("Расходы")
        ws2.merge_cells('A1:B1')
        ws2['A1'] = "Расходы по категориям"
        ws2['A1'].font = Font(bold=True, size=14)
        ws2['A1'].alignment = Alignment(horizontal='center')

        for col_idx, h in enumerate(['Категория', 'Сумма'], 1):
            cell = ws2.cell(row=3, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row_idx, (cat, total) in enumerate(sorted(data['expense_categories'].items()), 4):
            ws2.cell(row=row_idx, column=1, value=cat).border = thin_border
            val_c = ws2.cell(row=row_idx, column=2, value=total)
            val_c.border = thin_border
            val_c.alignment = Alignment(horizontal='right')
            if row_idx % 2 == 0:
                ws2.cell(row=row_idx, column=1).fill = alt_fill
                val_c.fill = alt_fill

        ws2.column_dimensions['A'].width = 35
        ws2.column_dimensions['B'].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_operational_excel(period='month', custom_start=None, custom_end=None):
    """
    Генерирует Excel с операционным отчётом для администратора (без финансов).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = _collect_operational_data(period, custom_start, custom_end)
    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    alt_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )

    # ── Лист 1: Заказы ──
    ws = wb.active
    ws.title = "Заказы"
    ws.merge_cells('A1:B1')
    ws['A1'] = data['title']
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    order_metrics = [
        ('Новые', str(data['new_orders'])),
        ('В работе', str(data['in_progress_orders'])),
        ('Готовые', str(data['ready_orders'])),
        ('Доставленные', str(data['delivered_orders'])),
        ('Просроченные', str(data['overdue_orders'])),
        ('Неоплаченные', str(data['unpaid_orders_count'])),
        ('Всего за период', str(data['total_orders'])),
    ]
    for col_idx, h in enumerate(['Статус', 'Количество'], 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    for row_idx, (label, val) in enumerate(order_metrics, 4):
        ws.cell(row=row_idx, column=1, value=label).border = thin_border
        vc = ws.cell(row=row_idx, column=2, value=val)
        vc.border = thin_border
        vc.alignment = Alignment(horizontal='right')
        if row_idx % 2 == 0:
            ws.cell(row=row_idx, column=1).fill = alt_fill
            vc.fill = alt_fill
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15

    # ── Лист 2: Склад ──
    ws2 = wb.create_sheet("Склад")
    ws2.merge_cells('A1:B1')
    ws2['A1'] = "Состояние склада"
    ws2['A1'].font = Font(bold=True, size=14)
    wh_metrics = [
        ('Материалов на складе', str(data['total_materials'])),
        ('С низким остатком', str(data['low_stock_materials'])),
        ('Готовой продукции', str(data['total_products'])),
    ]
    for col_idx, h in enumerate(['Показатель', 'Значение'], 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    for row_idx, (label, val) in enumerate(wh_metrics, 4):
        ws2.cell(row=row_idx, column=1, value=label).border = thin_border
        vc = ws2.cell(row=row_idx, column=2, value=val)
        vc.border = thin_border
        vc.alignment = Alignment(horizontal='right')
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 15

    # ── Лист 3: Задачи ──
    ws3 = wb.create_sheet("Задачи")
    ws3.merge_cells('A1:B1')
    ws3['A1'] = "Задачи"
    ws3['A1'].font = Font(bold=True, size=14)
    task_metrics = [
        ('Создано задач', str(data['tasks_created'])),
        ('Выполнено', str(data['tasks_completed'])),
    ]
    for col_idx, h in enumerate(['Показатель', 'Значение'], 1):
        cell = ws3.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    for row_idx, (label, val) in enumerate(task_metrics, 4):
        ws3.cell(row=row_idx, column=1, value=label).border = thin_border
        vc = ws3.cell(row=row_idx, column=2, value=val)
        vc.border = thin_border
        vc.alignment = Alignment(horizontal='right')
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 15

    # ── Лист 4: Работники ──
    if data.get('worker_stats'):
        ws4 = wb.create_sheet("Работники")
        ws4.merge_cells('A1:C1')
        ws4['A1'] = "Активность работников"
        ws4['A1'].font = Font(bold=True, size=14)
        for col_idx, h in enumerate(['Работник', 'Работ', 'Количество'], 1):
            cell = ws4.cell(row=3, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        for row_idx, w in enumerate(data['worker_stats'], 4):
            ws4.cell(row=row_idx, column=1, value=w['name']).border = thin_border
            ws4.cell(row=row_idx, column=2, value=w['works']).border = thin_border
            vc = ws4.cell(row=row_idx, column=3, value=w['quantity'])
            vc.border = thin_border
            vc.alignment = Alignment(horizontal='right')
        ws4.column_dimensions['A'].width = 20
        ws4.column_dimensions['B'].width = 12
        ws4.column_dimensions['C'].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
