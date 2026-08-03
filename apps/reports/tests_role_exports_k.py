"""
Отчёты для ролей: деньги только у владельца (баги тестировщиков 8 и 9).

8. «Отчёт по сотрудникам: логическая ошибка» — владелец видел в Excel/PDF
   выработку работников ровно как админ: только количества. Чтобы узнать
   начисленное за работы, приходилось собирать сумму из другой страницы.
   Теперь для владельца добавляется колонка «Начислено» (сумма labor_cost
   подтверждённых работ); админ денег по-прежнему не видит.

9. «В отчёте заказов нет колонки задолженности» — чтобы узнать, сколько
   клиент ещё должен по каждому заказу, нужно было считать вручную.
   Теперь у владельца в экспорте заказов есть колонка «Қарз»
   (total_amount - оплаченное); у администратора её нет — суммы ему
   не видны нигде в системе.
"""
import datetime
from decimal import Decimal

from django.test import TestCase
from django.utils import timezone
from rest_framework.test import APIClient

from apps.accounts.models import User
from apps.clients.models import Client, Payment
from apps.companies.models import Company
from apps.orders.models import Order
from apps.production.models import WorkRecord
from apps.warehouse.models import FinishedProduct

WORK_URL = '/api/v1/reports/export/work/'
ORDERS_URL = '/api/v1/reports/export/orders/'


def read_xlsx(content):
    import io
    from openpyxl import load_workbook
    return list(load_workbook(io.BytesIO(content)).active.iter_rows(values_only=True))


class RoleAwareWorkExportTests(TestCase):
    def setUp(self):
        self.company = Company.objects.create(name='RoleCo', is_active=True)
        self.owner = User.objects.create_user(username='role_owner', password='p',
                                              role=User.Role.OWNER, company=self.company)
        self.admin = User.objects.create_user(username='role_admin', password='p',
                                              role=User.Role.ADMIN, company=self.company)
        self.worker = User.objects.create_user(username='role_worker', password='p',
                                               role=User.Role.WORKER, company=self.company,
                                               full_name='Али')
        self.product = FinishedProduct.objects.create(
            company=self.company, name='Столешница', quantity=Decimal('10'), unit='dona')
        WorkRecord.objects.create(
            company=self.company, worker=self.worker, product=self.product,
            quantity=Decimal('4'), unit='dona', status=WorkRecord.WorkStatus.CONFIRMED,
            labor_cost=Decimal('600'), confirmed_at=timezone.now())
        self.api = APIClient()

    def test_owner_sees_labor_column(self):
        self.api.force_authenticate(self.owner)
        resp = self.api.get(WORK_URL)
        self.assertEqual(resp.status_code, 200)
        rows = read_xlsx(resp.content)
        self.assertEqual(len(rows[0]), 4, rows[0])
        self.assertIn('Али', rows[1])
        self.assertEqual(rows[1][3], Decimal('600'))

    def test_admin_does_not_see_money(self):
        self.api.force_authenticate(self.admin)
        resp = self.api.get(WORK_URL)
        rows = read_xlsx(resp.content)
        self.assertEqual(len(rows[0]), 3, rows[0])
        self.assertIn('Али', rows[1])
        self.assertEqual(len(rows[1]), 3)


class DebtColumnInOrdersExportTests(TestCase):
    def setUp(self):
        self.company = Company.objects.create(name='DebtCo', is_active=True)
        self.owner = User.objects.create_user(username='debt_owner', password='p',
                                              role=User.Role.OWNER, company=self.company)
        self.admin = User.objects.create_user(username='debt_admin', password='p',
                                              role=User.Role.ADMIN, company=self.company)
        self.cli = Client.objects.create(company=self.company, name='Клиент')
        self.product = FinishedProduct.objects.create(
            company=self.company, name='Столешница', quantity=Decimal('10'), unit='dona')
        self.order = Order.objects.create(
            company=self.company, client=self.cli, product=self.product,
            quantity=Decimal('1'), unit='dona', total_amount=Decimal('5000'),
            deadline=timezone.now() + datetime.timedelta(days=5))
        Payment.objects.create(company=self.company, client=self.cli, order=self.order,
                               amount=Decimal('2000'), payment_method='cash',
                               payment_date=timezone.now())
        self.api = APIClient()

    def test_owner_sees_debt_column_with_correct_value(self):
        self.api.force_authenticate(self.owner)
        resp = self.api.get(ORDERS_URL)
        self.assertEqual(resp.status_code, 200)
        rows = read_xlsx(resp.content)
        self.assertEqual(len(rows[0]), 8, rows[0])
        self.assertEqual(rows[1][0], self.order.id)
        self.assertEqual(rows[1][7], Decimal('3000'))  # 5000 - 2000

    def test_paid_order_has_zero_debt(self):
        Payment.objects.create(company=self.company, client=self.cli, order=self.order,
                               amount=Decimal('3000'), payment_method='cash',
                               payment_date=timezone.now())
        self.api.force_authenticate(self.owner)
        rows = read_xlsx(self.api.get(ORDERS_URL).content)
        self.assertEqual(rows[1][7], 0)

    def test_custom_product_order_is_exported_with_debt(self):
        Order.objects.create(
            company=self.company, client=self.cli, custom_product_name='По эскизу',
            quantity=Decimal('2'), unit='dona', total_amount=Decimal('1000'),
            deadline=timezone.now() + datetime.timedelta(days=5))
        self.api.force_authenticate(self.owner)
        rows = read_xlsx(self.api.get(ORDERS_URL).content)
        # Экспорт идёт в порядке модели: -created_at, свежий заказ первый.
        self.assertEqual(rows[1][2], 'По эскизу')
        self.assertEqual(rows[1][7], Decimal('1000'))
        self.assertEqual(rows[2][7], Decimal('3000'))

    def test_admin_export_has_no_money_columns(self):
        self.api.force_authenticate(self.admin)
        rows = read_xlsx(self.api.get(ORDERS_URL).content)
        self.assertEqual(len(rows[0]), 7, rows[0])
        self.assertEqual(len(rows[1]), 7)
